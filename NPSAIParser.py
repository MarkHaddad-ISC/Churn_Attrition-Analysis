from openpyxl import load_workbook
from Prompts.NPSPrompt import *
from Functions.vanillaAI import *
from Functions.OutputCleansers import *
from Dictionaries.numberToReasonLostMapping import *
import math, time

numberOfOutputsFromAI = 2
start_time = time.time()
# Define the file path and sheet name
file_path = '/Users/haddadm/CursorXopenPYXL/RawInputs/RawInput_NPSData.xlsx'
sheet_name = 'CleanData'

# Batch size for incremental saving
batch_size = 10

starting_row = 2 #change to actual starting row value if it errors out, no offset neccesary, make this number the exact number of the row you want to start anaylsis on


# Load the workbook and select the sheet
workbook = load_workbook(filename=file_path, data_only=True)
sheet = workbook[sheet_name]

# Determine final row (excluding header), the value of this line should be equal to the number of the last row of excel data offset by a value of 1
FinalRowPlus1 = len([cell.value for cell in sheet['D'] if cell.value is not None]) + 1
#FinalRowPlus1 = 13  # override for testing or partial runs

print(FinalRowPlus1)

# Extract entries from column M (excluding the header)
column_m_values = [sheet[f'E{row}'].value for row in range(starting_row, FinalRowPlus1)]
entries = [str(value) for value in column_m_values]

numberOfBatches = math.ceil((FinalRowPlus1 - starting_row) / batch_size)
total_entries = len(entries)

print(total_entries)
print("")
print(f"Performing analysis on {numberOfBatches} batches of size {batch_size} for a total of {FinalRowPlus1 - starting_row} rows")
print("")

for batch_start in range(0, total_entries, batch_size):
    batch_end = min(batch_start + batch_size, total_entries)
    batch_entries = entries[batch_start:batch_end]

    strToList = '['
    counter = 0

    for item in batch_entries:
        while True:
            singularLineAnalysis = clean_string(vanillaChatBot(NPSprompt + item, systemPythonPrompt, 'gpt-4o'))
            productReasonSummary = extract_bracketed_value(singularLineAnalysis)
            if productReasonSummary == False:
                print('redo!')
                continue
            else:
                strToList += productReasonSummary + ','
                counter += 1
                print(productReasonSummary)
                print('appending entry #' + str(batch_start + counter))
                break

    strToList = strToList[:-1] + ']'
    analyzed_batch = convert_to_dict(strToList)

    # Add primary reason mapping
    for key, value in analyzed_batch.items():
        secondary_code = value[1]
        primary_reason = reasonSecondaryToPrimary.get(secondary_code, 'Unknown Reason')
        value.insert(2, primary_reason)

    # Add secondary reason mapping
    for key, value in analyzed_batch.items():
        risk_code = value[1]
        if risk_code in retention_risk_reasons:
            value[1] = retention_risk_reasons[risk_code]

    # Write to Excel — offset row number by batch_start
    for local_row_num, values in analyzed_batch.items():
        #global_row_num = batch_start + local_row_num + 1  # +1 to account for header
        global_row_num = starting_row + batch_start + local_row_num - 1  # -1 to account for header
        for col_num, value in enumerate(values, start=10):  # Start writing at column Z (26)
            workbook.active.cell(row=global_row_num, column=col_num).value = value

    workbook.save(file_path)
    print(f"✅ Excel Sheet Lines {starting_row + batch_start} to {starting_row + batch_end - 1} have been saved successfully.")

end_time = time.time()
print("✅ All batches processed and saved.")
print("")
print(f"\nTotal execution time: {end_time - start_time:.2f} seconds")