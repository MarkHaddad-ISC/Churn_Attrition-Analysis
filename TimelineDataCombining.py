# RUN THIS SCRIPT SECOND
# take the SF data and cleans it by doing the following:
#   Include if attrition — Only if account ID shows attrition in GA export
#   Before attrition date — Use meeting notes only before GA date
#   Must contain keywords — Include if it has: cancellation, renew, terminate, downgrade, etc.
#   Exclude promo terms — Skip if it has: Register Now, Webinar, Join Us, etc.
#   Remove duplicates — Exclude repeated entries
from openpyxl import load_workbook, Workbook
from Functions.DateFilteringForTimelineData import *
import pandas as pd

# This is the salesforce export, it was very large (>400,000) so it had to be split up between multiple excel documents
file1_path = '/Users/haddadm/CursorXopenPYXL/RawInputs/RawInput_Oct_Dec_2024_Timeline2.xlsx'
file2_path = '/Users/haddadm/CursorXopenPYXL/RawInputs/RawInput_Jan_Sept_2025_Timeline2.xlsx'
file3_path = '/Users/haddadm/CursorXopenPYXL/RawInputs/RawInput_Sept_Oct_2025_Timeline2.xlsx'




# REPLACE THIS WITH ARR DATA PATH
arr_file_path = '/Users/haddadm/CursorXopenPYXL/RawInputs/RawInput_RMS_Timeline_Only.xlsx'
output_path = '/Users/haddadm/CursorXopenPYXL/Outputs/OUTPUT_TimelineData.xlsx'

sheet1_name = sheet2_name = sheet3_name = 'RawData'


arr_sheet_name = 'ARR_RMS'
output_sheet_name = 'RMSOnly'





excluded_substrings = [
    "Register Now",
    "NPS Survey",
    "Your Invitation",
    "Join Us",
    "Don't Miss Out",
    "Webinar"
]


included_substrings = [
"cancellation",
"cancel",
"renewal",
"renew",
"terminate",
"termination",
"downgrade",
"reduce",
"lower",
"unsubscribe",
"subscription",
"license",
"unsatisfied", 
"disappointed", 
"lack of value", 
"unhappy", 
"frustrated",
"un valuable", 
"not valuable", 
"not useful", 
"low usage", 
"limited usage",
"low value", 
"limited value", 
"poorly", 
"upset", 
"irritated"
]



headers = [
    'Description', 'WhatId', 'Status', 'Subject', 'ActivityDate',
    'AccountID', 'Segment', 'Region', 'Account_Name', 'WhatID Object',
    'Reason Lost', 'Reason Lost Notes', 'Reason Lost Secondary'
]


# Define the columns to retrieve (1-based indexing)
columns_to_extract = [2, 7, 8, 9, 10, 12, 19, 20, 22, 24, 25, 26, 27]

# Step 1.1: Extract accountIDs from column TBD of the ARR data
ga_df = pd.read_excel(arr_file_path, sheet_name=arr_sheet_name, engine='openpyxl')
ARR_AccountIDs = ga_df.iloc[1:, 1].dropna().astype(str).tolist()
print(ARR_AccountIDs[1])
# Step 1.2: Extract Dates from column TBD of the ARR data
#ARR_Dates = ga_df.iloc[1:, 1].dropna().astype(str).tolist()





# Step 2: Function to extract and filter relevant data from a sheet
def extract_filtered_data(file_path, sheet_name, filter_values):
    wb = load_workbook(filename=file_path, data_only=True)
    ws = wb[sheet_name]
    filtered_data = []
    for row in ws.iter_rows(min_row=2):  # Skip the first row
        selected_row = [row[i-1].value for i in columns_to_extract]
        if selected_row[5] is not None and str(selected_row[5]) in filter_values:
            filtered_data.append(selected_row)
    return filtered_data

# Extract and filter data from both sheets


filtered_data1 = extract_filtered_data(file1_path, sheet1_name, ARR_AccountIDs)
filtered_data2 = extract_filtered_data(file2_path, sheet2_name, ARR_AccountIDs)
filtered_data3 = extract_filtered_data(file3_path, sheet3_name, ARR_AccountIDs)

# Combine the filtered data filtered_data3 is cronologically the earliest data field
combined_filtered_data =  filtered_data1 + filtered_data2 + filtered_data3



# this filter looks to see if our included_substrings are included in either the decription or email topic
descriptonSet = set()
for i in range(len(combined_filtered_data) - 1, -1, -1):
    if any(substring in str(combined_filtered_data[i][3]).lower() or substring in str(combined_filtered_data[i][0]).lower() for substring in included_substrings):
        #this removes duplicate description entries
        if combined_filtered_data[i][0] in descriptonSet:
            del combined_filtered_data[i]
        else:    
            descriptonSet.add(combined_filtered_data[i][0])
        continue
    else:
        del combined_filtered_data[i]

print(len(combined_filtered_data))
"""
for i in range(len(combined_filtered_data)):
    if str(combined_filtered_data[i][4])[5:7] == "11" or str(combined_filtered_data[i][4])[5:7] == "12":
        print("AccountID: " + str(combined_filtered_data[i][5]) + ",  Month: " + str(combined_filtered_data[i][4])[5:7])
"""
for i in range(len(combined_filtered_data) - 1, -1, -1):
    if any(substring in str(combined_filtered_data[i][3]) for substring in excluded_substrings):
        del combined_filtered_data[i]
print(len(combined_filtered_data))

#this is the code for date filtering
"""
latestDatePerAccountDict = build_max_value_dict(ARR_AccountIDs, ARR_Dates)
final_Filtered_Data = filter_rows_by_dict(combined_filtered_data, latestDatePerAccountDict)
"""

print(len(combined_filtered_data))
# Create a new workbook and write the filtered data
output_wb = Workbook()
output_ws = output_wb.active
output_ws.title = output_sheet_name

# Insert headers in the second row
output_ws.append(list(headers))

#uncomment this as well and delete the for loop before to do this
"""
for row in final_Filtered_Data:
    output_ws.append(row)
"""
for row in combined_filtered_data:
    output_ws.append(row)
    
  

# Save the output workbook
output_wb.save(output_path)

print(f"Filtered and combined data has been saved to '{output_path}' in the sheet '{output_sheet_name}'.")


