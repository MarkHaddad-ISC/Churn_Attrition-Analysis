from openpyxl import load_workbook
from Prompts.GSPrompt import *
from Functions.vanillaAI import *
from Functions.OutputCleansers import *
from Dictionaries.numberToReasonLostMapping import *
import math, time, requests



start_time = time.time()
#caffeinate_process = subprocess.Popen(['caffeinate', '-i'])
# Define the file path and sheet name
file_path = '/Users/haddadm/CursorXopenPYXL/RawInputs/RawInput_GS_CSM_Qualitative.xlsx'
sheet_name = 'RawData'
GPTModel = 'gpt-4o'
#GPTModel = 'gpt-5'

numberOfOutputsFromAI = 2

# Batch size for incremental saving, Value should be no less than 50, theoretically shouldnt matter with new error handeling, batch size is inversely proportinal to runtime
batch_size = 5
starting_row = 2 #change to actual starting row value if it errors out, no offset neccesary, make this number the exact number of the row you want to start anaylsis on, row 1 is the header row no naturally we start at 2


# Load the workbook and select the sheet
workbook = load_workbook(filename=file_path, data_only=True)
sheet = workbook[sheet_name]

# Determine final row (excluding header)
FinalRowPlus1 = len([cell.value for cell in sheet['N'] if cell.value is not None]) + 1

print(FinalRowPlus1)





commentsValues = [sheet[f'A{row}'].value for row in range(starting_row, FinalRowPlus1)]
commentsEntries = [str(value) for value in commentsValues]


healthValues = [sheet[f'J{row}'].value for row in range(starting_row, FinalRowPlus1)]
healthEntries = [str(value) for value in healthValues]

# Combine each comment with its corresponding health score
entries = [f"Meeting Notes: {comment} || Account Health: {health}" 
           for comment, health in zip(commentsEntries, healthEntries)]

print(entries[1])

numberOfBatches = math.ceil((FinalRowPlus1 - starting_row) / batch_size)
total_entries = len(entries)

print(total_entries)
print("")
print(f"Performing analysis on {numberOfBatches} batches of size {batch_size} for a total of {FinalRowPlus1 - starting_row} rows")
print("")

MAX_RETRIES = 5
RETRY_DELAY = 15  # seconds

# Exception counters
connection_error_count = 0
generic_error_count = 0

for batch_start in range(0, total_entries, batch_size):
    batch_end = min(batch_start + batch_size, total_entries)
    batch_entries = entries[batch_start:batch_end]

    strToList = '['
    counter = 0

    for item in batch_entries:
        for attempt in range(MAX_RETRIES):
            try:
                singularLineAnalysis = clean_string(vanillaChatBot(promptCSMData + item, systemPythonPrompt, GPTModel))
                productReasonSummary = extract_bracketed_value(singularLineAnalysis, numberOfOutputsFromAI)
                if productReasonSummary:
                    strToList += productReasonSummary + ','
                    counter += 1
                    print(productReasonSummary)
                    print('appending line #' + str(starting_row + batch_start + counter - 1))
                    break
                else:
                    print('redo due to parsing failure!')
                    time.sleep(RETRY_DELAY)
            except requests.exceptions.ConnectionError as e:
                connection_error_count += 1
                print(f"[{starting_row + batch_start + counter - 1}] Connection error on attempt {attempt+1}: {e}")
                time.sleep(RETRY_DELAY)
            except Exception as e:
                generic_error_count += 1
                print(f"[{starting_row + batch_start + counter - 1}] Unexpected error on attempt {attempt+1}: {e}")
                time.sleep(RETRY_DELAY)
        else:
            print(f"❌ Failed after {MAX_RETRIES} attempts for entry #{starting_row + batch_start + counter - 1}")
            strToList += '"ERROR",'
            counter += 1

    strToList = strToList[:-1] + ']'
    analyzed_batch = convert_to_dict(strToList)

    # Add primary reason mapping
    for key, value in analyzed_batch.items():
        secondary_code = value[1]
        primary_reason = reasonSecondaryToPrimary.get(secondary_code, 'Unknown Reason')
        value.insert(2, primary_reason)

    # Add secondary reason mapping
    for key, value in analyzed_batch.items():
        risk_code = value[1]
        if risk_code in retention_risk_reasons:
            value[1] = retention_risk_reasons[risk_code]

    # Write to Excel
    for local_row_num, values in analyzed_batch.items():
        global_row_num = starting_row + batch_start + local_row_num - 1
        for col_num, value in enumerate(values, start=17):  # Start at column Q
            sheet.cell(row=global_row_num, column=col_num).value = value

    workbook.save(file_path)
    print(f"✅ Excel Sheet Lines {starting_row + batch_start} to {starting_row + batch_end - 1} have been saved successfully.")

end_time = time.time()
hoursSpent = (end_time - start_time) / 3600
minutesRemainder = (hoursSpent - int(hoursSpent))*60
print("✅✅✅ All batches processed and saved. ✅✅✅")
print("")
print(f"Total execution time: {end_time - start_time:.2f} seconds")
print("")
print(f"Total execution time: {int(hoursSpent)} hours and {int(minutesRemainder)} minutes")
print("")
print(f"🔁 Connection errors encountered: {connection_error_count}")
print(f"⚠️ Other exceptions encountered: {generic_error_count}")

#caffeinate_process.terminate()
