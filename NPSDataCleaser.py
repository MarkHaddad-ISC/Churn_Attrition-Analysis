import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
from typing import Dict, List, Optional, Any


def load_nps_data(
    file_path: str,
    sheet_name: str = "2024 Q4 to 2025 Q3 NPS Data",
    column_mapping: Optional[Dict[int, str]] = None,
    start_row: int = 2,
    skip_empty_rows: bool = True
) -> pd.DataFrame:
    """
    Load NPS data from an Excel file using openpyxl.
    
    Args:
        file_path (str): Path to the Excel file (e.g., "RawInput_NPSData.xlsx")
        sheet_name (str): Name of the sheet to load
        column_mapping (Dict[int, str]): Dictionary mapping column numbers (1-based) to column names.
            Default mapping:
            {
                1: "Year",
                2: "Quarter", 
                3: "NPS_Score",
                4: "NPS_Group",
                5: "NPS_Comments",
                6: "Survey_Channel",
                7: "Account_ID",
                8: "Old_Macro_Product",
                9: "New_Macro_Product"
            }
        start_row (int): Row number to start reading data from (1-based, default is 2 to skip header)
        skip_empty_rows (bool): Whether to skip completely empty rows
        
    Returns:
        pd.DataFrame: DataFrame containing the NPS data with specified columns
    """
    
    # Default column mapping
    if column_mapping is None:
        column_mapping = {
            1: "Year",
            2: "Quarter", 
            3: "NPS_Score",
            4: "NPS_Group",
            5: "NPS_Comments",
            6: "Survey_Channel",
            7: "Account_ID",
            8: "Old_Macro_Product",
            9: "New_Macro_Product"
        }
    
    try:
        # Load the workbook
        print(f"Loading workbook: {file_path}")
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        # Check if sheet exists
        if sheet_name not in workbook.sheetnames:
            print(f"Error: Sheet '{sheet_name}' not found.")
            print(f"Available sheets: {', '.join(workbook.sheetnames)}")
            return pd.DataFrame()
        
        # Get the worksheet
        worksheet = workbook[sheet_name]
        print(f"Loading sheet: {sheet_name}")
        
        # Determine max row and column with data
        max_row = worksheet.max_row
        max_col = max(column_mapping.keys())
        
        print(f"Sheet dimensions: {max_row} rows x {max_col} columns")
        
        # Initialize data dictionary
        data = {col_name: [] for col_name in column_mapping.values()}
        
        # Read data from specified columns
        rows_read = 0
        for row_idx in range(start_row, max_row + 1):
            # Check if row is empty (if skip_empty_rows is True)
            if skip_empty_rows:
                row_values = [worksheet.cell(row_idx, col_num).value 
                             for col_num in column_mapping.keys()]
                if all(v is None or (isinstance(v, str) and v.strip() == "") 
                      for v in row_values):
                    continue
            
            # Read each column for this row
            for col_num, col_name in column_mapping.items():
                cell_value = worksheet.cell(row_idx, col_num).value
                data[col_name].append(cell_value)
            
            rows_read += 1
        
        # Close the workbook
        workbook.close()
        
        # Create DataFrame
        df = pd.DataFrame(data)
        
        print(f"✓ Successfully loaded {rows_read} rows of data")
        print(f"Columns: {', '.join(df.columns)}")
        
        return df
        
    except FileNotFoundError:
        print(f"Error: File '{file_path}' not found.")
        return pd.DataFrame()
    
    except Exception as e:
        print(f"Error loading NPS data: {str(e)}")
        return pd.DataFrame()


def get_sheet_info(file_path: str) -> None:
    """
    Display information about all sheets in the Excel file.
    
    Args:
        file_path (str): Path to the Excel file
    """
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        print(f"\nWorkbook: {file_path}")
        print("=" * 60)
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            print(f"\nSheet: {sheet_name}")
            print(f"  Dimensions: {worksheet.max_row} rows x {worksheet.max_column} columns")
            
            # Display first row (header)
            header_row = []
            for col_idx in range(1, min(worksheet.max_column + 1, 20)):  # Limit to 20 columns
                cell_value = worksheet.cell(1, col_idx).value
                if cell_value:
                    header_row.append(f"Col{col_idx}: {cell_value}")
            
            if header_row:
                print(f"  Headers: {', '.join(header_row)}")
        
        workbook.close()
        
    except Exception as e:
        print(f"Error reading workbook info: {str(e)}")


def filter_nps_data(
    df: pd.DataFrame,
    min_comment_chars: int = 15,
    min_comment_words: int = 4,
    exclude_nps_groups: Optional[List[str]] = None,
    exclude_old_products: Optional[List[str]] = None,
    verbose: bool = True
) -> pd.DataFrame:
    """
    Filter NPS data based on specified criteria.
    
    Args:
        df (pd.DataFrame): Input DataFrame with NPS data
        min_comment_chars (int): Minimum character length for NPS_Comments (default: 15)
        min_comment_words (int): Minimum word count for NPS_Comments (default: 4)
        exclude_nps_groups (List[str]): List of NPS_Group values to exclude (default: ["Promoter"])
        exclude_old_products (List[str]): List of Old_Macro_Product values to exclude (default: ["Learning Solutions"])
        verbose (bool): Print filtering statistics (default: True)
        
    Returns:
        pd.DataFrame: Filtered DataFrame
    """
    
    if df.empty:
        print("Warning: Input DataFrame is empty.")
        return df
    
    # Set default exclusions
    if exclude_nps_groups is None:
        exclude_nps_groups = ["Promoter"]
    
    if exclude_old_products is None:
        exclude_old_products = ["Learning Solutions"]
    
    original_count = len(df)
    df_filtered = df.copy()
    
    if verbose:
        print(f"\nFiltering NPS Data...")
        print("=" * 60)
        print(f"Original row count: {original_count}")
    
    # Filter 1: Remove rows with short or few-word comments
    if 'NPS_Comments' in df_filtered.columns:
        # Handle NaN values - treat them as empty strings
        df_filtered['NPS_Comments'] = df_filtered['NPS_Comments'].fillna('')
        
        # Convert to string (in case of numeric values)
        df_filtered['NPS_Comments'] = df_filtered['NPS_Comments'].astype(str)
        
        # Create masks for filtering
        char_mask = df_filtered['NPS_Comments'].str.len() >= min_comment_chars
        word_mask = df_filtered['NPS_Comments'].str.split().str.len() >= min_comment_words
        
        # Keep rows that meet BOTH criteria (at least X chars AND at least Y words)
        comment_mask = char_mask & word_mask
        
        rows_before = len(df_filtered)
        df_filtered = df_filtered[comment_mask]
        rows_removed = rows_before - len(df_filtered)
        
        if verbose:
            print(f"  ✓ Removed {rows_removed} rows with comments < {min_comment_chars} chars or < {min_comment_words} words")
    else:
        if verbose:
            print("  ⚠ Warning: 'NPS_Comments' column not found, skipping comment filter")
    
    # Filter 2: Remove rows where NPS_Group is in exclusion list
    if 'NPS_Group' in df_filtered.columns:
        rows_before = len(df_filtered)
        df_filtered = df_filtered[~df_filtered['NPS_Group'].isin(exclude_nps_groups)]
        rows_removed = rows_before - len(df_filtered)
        
        if verbose:
            print(f"  ✓ Removed {rows_removed} rows where NPS_Group in {exclude_nps_groups}")
    else:
        if verbose:
            print("  ⚠ Warning: 'NPS_Group' column not found, skipping NPS group filter")
    
    # Filter 3: Remove rows where Old_Macro_Product is in exclusion list
    if 'Old_Macro_Product' in df_filtered.columns:
        rows_before = len(df_filtered)
        df_filtered = df_filtered[~df_filtered['Old_Macro_Product'].isin(exclude_old_products)]
        rows_removed = rows_before - len(df_filtered)
        
        if verbose:
            print(f"  ✓ Removed {rows_removed} rows where Old_Macro_Product in {exclude_old_products}")
    else:
        if verbose:
            print("  ⚠ Warning: 'Old_Macro_Product' column not found, skipping product filter")
    
    # Filter 4: Remove rows where Account_ID is null or missing
    if 'Account_ID' in df_filtered.columns:
        rows_before = len(df_filtered)
        # Remove rows where Account_ID is null, empty string, or whitespace only
        df_filtered = df_filtered[df_filtered['Account_ID'].notna()]
        df_filtered = df_filtered[df_filtered['Account_ID'].astype(str).str.strip() != '']
        rows_removed = rows_before - len(df_filtered)
        
        if verbose:
            print(f"  ✓ Removed {rows_removed} rows where Account_ID is null or missing")
    else:
        if verbose:
            print("  ⚠ Warning: 'Account_ID' column not found, skipping Account_ID filter")
    
    # Reset index
    df_filtered = df_filtered.reset_index(drop=True)
    
    final_count = len(df_filtered)
    total_removed = original_count - final_count
    
    if verbose:
        print(f"\nFinal row count: {final_count}")
        print(f"Total rows removed: {total_removed} ({total_removed/original_count*100:.1f}%)")
        print(f"Retention rate: {final_count/original_count*100:.1f}%")
        print("=" * 60)
    
    return df_filtered


def write_clean_data(
    df: pd.DataFrame,
    file_path: str,
    sheet_name: str = "CleanData",
    include_index: bool = False,
    overwrite_sheet: bool = True
) -> bool:
    """
    Write the cleaned DataFrame back to the Excel file in a new worksheet.
    
    Args:
        df (pd.DataFrame): The cleaned DataFrame to write
        file_path (str): Path to the Excel file
        sheet_name (str): Name of the sheet to create/write to (default: "CleanData")
        include_index (bool): Whether to include the DataFrame index (default: False)
        overwrite_sheet (bool): If True, overwrites existing sheet with same name (default: True)
        
    Returns:
        bool: True if successful, False otherwise
    """
    
    if df.empty:
        print("Warning: Cannot write empty DataFrame.")
        return False
    
    try:
        # Load the existing workbook
        print(f"Opening workbook: {file_path}")
        workbook = openpyxl.load_workbook(file_path)
        
        # Check if sheet already exists
        if sheet_name in workbook.sheetnames:
            if overwrite_sheet:
                print(f"Sheet '{sheet_name}' already exists. Removing old sheet...")
                del workbook[sheet_name]
            else:
                print(f"Error: Sheet '{sheet_name}' already exists and overwrite_sheet=False")
                workbook.close()
                return False
        
        # Create new worksheet
        worksheet = workbook.create_sheet(sheet_name)
        print(f"Creating new sheet: {sheet_name}")
        
        # Write column headers
        columns = list(df.columns)
        if include_index:
            columns = [df.index.name or 'Index'] + columns
        
        for col_idx, col_name in enumerate(columns, start=1):
            worksheet.cell(1, col_idx, col_name)
        
        # Write data rows
        for row_idx, (_, row_data) in enumerate(df.iterrows(), start=2):
            col_offset = 0
            
            # Write index if requested
            if include_index:
                worksheet.cell(row_idx, 1, row_data.name)
                col_offset = 1
            
            # Write row data
            for col_idx, value in enumerate(row_data, start=1 + col_offset):
                worksheet.cell(row_idx, col_idx, value)
        
        # Save the workbook
        workbook.save(file_path)
        workbook.close()
        
        print(f"✓ Successfully wrote {len(df)} rows to sheet '{sheet_name}'")
        print(f"✓ File saved: {file_path}")
        return True
        
    except FileNotFoundError:
        print(f"Error: File '{file_path}' not found.")
        return False
    
    except PermissionError:
        print(f"Error: Permission denied. Make sure the file '{file_path}' is not open in another program.")
        return False
    
    except Exception as e:
        print(f"Error writing clean data: {str(e)}")
        return False


def preview_data(df: pd.DataFrame, num_rows: int = 5) -> None:
    """
    Display a preview of the DataFrame.
    
    Args:
        df (pd.DataFrame): The DataFrame to preview
        num_rows (int): Number of rows to display
    """
    if df.empty:
        print("DataFrame is empty.")
        return
    
    print(f"\nData Preview (first {num_rows} rows):")
    print("=" * 80)
    print(df.head(num_rows).to_string())
    print(f"\nTotal rows: {len(df)}")
    print(f"Total columns: {len(df.columns)}")
    
    # Display data types
    print("\nColumn Data Types:")
    for col, dtype in df.dtypes.items():
        print(f"  {col}: {dtype}")
    
    # Display null counts
    null_counts = df.isnull().sum()
    if null_counts.sum() > 0:
        print("\nNull/Missing Values:")
        for col, count in null_counts.items():
            if count > 0:
                print(f"  {col}: {count} ({count/len(df)*100:.1f}%)")


# Example usage
if __name__ == "__main__":
    # Example 1: Load data with default column mapping
    file_path = "RawInput_NPSData.xlsx"
    
    # First, check what's in the file
    print("Inspecting workbook structure...")
    get_sheet_info(file_path)
    
    print("\n" + "=" * 80)
    print("Loading NPS data with default mapping...")
    print("=" * 80)
    
    # Load the data
    df = load_nps_data(file_path)
    
    # Preview the data
    if not df.empty:
        preview_data(df)
        
        # Apply filters
        df_filtered = filter_nps_data(df)
        
        # Preview filtered data
        print("\n" + "=" * 80)
        print("Filtered Data:")
        print("=" * 80)
        preview_data(df_filtered)
        
        # Write cleaned data back to Excel
        print("\n" + "=" * 80)
        print("Writing cleaned data to Excel...")
        print("=" * 80)
        write_clean_data(df_filtered, file_path, sheet_name="CleanData")
    
    # Example 2: Load data with custom column mapping
    # custom_mapping = {
    #     1: "Year",
    #     2: "Quarter",
    #     3: "Score",
    #     5: "Comments",
    #     7: "Account"
    # }
    # df_custom = load_nps_data(file_path, column_mapping=custom_mapping)
    
    # Example 3: Apply custom filters
    # df_custom_filtered = filter_nps_data(
    #     df,
    #     min_comment_chars=20,
    #     min_comment_words=5,
    #     exclude_nps_groups=["Promoter", "Passive"],
    #     exclude_old_products=["Learning Solutions", "Orbis"]
    # )

