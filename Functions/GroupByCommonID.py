from Functions.vanillaAI import *
from Prompts.SystemPrompt import *
from openpyxl import load_workbook
from openpyxl import workbook
from Functions.OutputCleansers import *
from Dictionaries.numberToReasonLostMapping import *
import math
import time
import ast

def group_rows_by_contract_id(sheet, start_row, end_row, key_col='A', group_col='B', data_cols=None):
    """
    Groups rows by Contract_ID and creates a nested dictionary structure.
    
    Args:
        sheet: openpyxl worksheet object
        start_row: starting row for processing
        end_row: ending row for processing
        key_col: column letter to use as the inner dictionary key (e.g., 'A')
        group_col: column letter for grouping (e.g., 'B' for Contract_ID)
        data_cols: list of column letters to include in the value list (e.g., ['C','D','E','F'])
    
    Returns:
        dict: {contract_id: {row_key: [values from data_cols]}}
    """
    if data_cols is None:
        raise ValueError("You must provide a list of data columns to include.")
    
    grouped_data = {}
    
    for row in range(start_row, end_row):
        contract_id = sheet[f'{group_col}{row}'].value
        row_key = sheet[f'{key_col}{row}'].value
        
        
        if contract_id and row_key:
            if contract_id not in grouped_data:
                grouped_data[contract_id] = {}
            
            # Collect values from specified columns
            row_values = [sheet[f'{col}{row}'].value for col in data_cols]
            grouped_data[contract_id][row_key] = row_values
    
    return grouped_data



def create_batches(grouped_entries, batch_size=50):
    """
    Splits grouped Contract_ID data into batches.
    
    Args:
        grouped_entries (dict): {contract_id: {row_key: [values]}}
        batch_size (int): Number of Contract_ID groups per batch.
    
    Returns:
        list of dicts: Each dict is a batch of contract_id -> rows_dict
    """
    unique_ids = list(grouped_entries.keys())
    number_of_batches = math.ceil(len(unique_ids) / batch_size)
    
    batches = []
    for i in range(0, len(unique_ids), batch_size):
        batch_ids = unique_ids[i:i + batch_size]
        batch_dict = {cid: grouped_entries[cid] for cid in batch_ids}
        batches.append(batch_dict)
    
    print(f"✅ Created {number_of_batches} batches of size {batch_size}")
    return batches



def process_batches_and_writeV3(workbook, file_path, sheet, batches, ai_model, prompt,
                                save_interval=5, start_col=27,
                                numberOfThemesToExtract=3):
    """
    Processes batches, runs AI analysis, cleans output, maps codes, and writes to Excel.
    Uses extract_bracketed_value and convert_to_dict for structured parsing.
    """
    ai_call_count = 0
    MAX_RETRIES = 5
    RETRY_DELAY = 10

    for batch in batches:
        for contract_id, rows_dict in batch.items():
            combined_text = "\n".join([f"{k}: {v}" for k, v in rows_dict.items()])

            # Retry AI call
            ai_output_clean = None
            for attempt in range(MAX_RETRIES):
                try:
                    print("combined_text " + combined_text)
                    raw_output = vanillaChatBot(prompt + combined_text, systemPythonPrompt, ai_model)
                    cleaned_output = clean_string(raw_output)

                    # Extract bracketed value if valid
                    bracketed_value = extract_bracketed_value(cleaned_output, numberOfThemesToExtract)
                    if bracketed_value:
                        ai_output_clean = bracketed_value
                        break
                    else:
                        print(f"Parsing failed on attempt {attempt+1}, retrying...")
                        time.sleep(RETRY_DELAY)
                except Exception as e:
                    print(f"Error on attempt {attempt+1}: {e}")
                    time.sleep(RETRY_DELAY)

            if not ai_output_clean:
                print(f"❌ Failed after {MAX_RETRIES} attempts for Contract_ID {contract_id}")
                ai_output_clean = "[['ERROR', 0, 'UNKNOWN']]"

            # Convert AI output to dict of lists
            analyzed_dict = convert_to_dict(ai_output_clean)  # Example: {1: ['Reason', 15, 'Extra']}
            print(analyzed_dict)

            # Write enriched output for each row in this Contract_ID group
            for local_key, values in analyzed_dict.items():
                # Extract values safely

                # Extract values from analyzed_dict
                reason_text = analyzed_dict.get(1, "Unknown")
                secondary_code = analyzed_dict.get(2, 0)
                extra_value = analyzed_dict.get(3, "")

                # Map secondary code
                secondary_reason = retention_risk_reasons.get(secondary_code, "Unknown Risk")
                primary_reason = reasonSecondaryToPrimary.get(secondary_code, "Unknown Reason")

                # Prepare final values for writing
                final_values = [reason_text, secondary_reason, primary_reason, extra_value]



                for row_key in rows_dict.keys():
                    excel_row = int(row_key)
                    for col_offset, value in enumerate(final_values):
                            sheet.cell(row=excel_row, column=start_col + col_offset).value = value

                """
                # Determine Excel row (map local_key to actual row in rows_dict)
                row_keys = list(rows_dict.keys())
                if local_key - 1 < len(row_keys):
                    row_key = row_keys[local_key - 1]
                else:
                    row_key = row_keys[-1]  # fallback to last row

                excel_row = int(row_key) + 1  # assuming numeric IDs
                for col_offset, value in enumerate(final_values):
                    sheet.cell(row=excel_row, column=start_col + col_offset).value = value
                """

            ai_call_count += 1
            if ai_call_count % save_interval == 0:
                workbook.save(file_path)
                print(f"💾 Saved after {ai_call_count} AI calls.")

    workbook.save(file_path)
    print("✅ All batches processed and saved.")