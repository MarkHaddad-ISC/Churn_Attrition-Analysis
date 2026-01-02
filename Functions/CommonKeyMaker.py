from openpyxl import load_workbook
import time
import sys
import os
# Add parent directory to path to import from Dictionaries
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from Dictionaries.slideCategories import slideCategories


def concatenate_columns_and_insert_to_column_a(file_path, sheet_name, col1, col2, start_row=1, save_path=None):
    """
    Concatenates two columns with " - " separator and inserts the result into column A.
    All operations are performed in memory after loading for better performance.
    
    Parameters:
    -----------
    file_path : str
        Path to the Excel file
    sheet_name : str
        Name of the sheet to process
    col1 : str or int
        First column to concatenate (e.g., 'B' or 2)
    col2 : str or int
        Second column to concatenate (e.g., 'C' or 3)
    start_row : int, optional
        Row to start processing from (default: 1, header row is typically 1)
    save_path : str, optional
        Path to save the modified file. If None, overwrites the original file.
    
    Returns:
    --------
    float: Time taken in seconds
    """
    start_time = time.time()
    
    # Load the workbook
    load_start = time.time()
    workbook = load_workbook(filename=file_path, data_only=True)
    sheet = workbook[sheet_name]
    load_time = time.time() - load_start
    print(f"⏱️  Load time: {load_time:.2f} seconds")
    
    # Convert column letters to numbers if needed (0-indexed for list access)
    if isinstance(col1, str):
        col1_num = ord(col1.upper()) - ord('A')
    else:
        col1_num = col1 - 1
    
    if isinstance(col2, str):
        col2_num = ord(col2.upper()) - ord('A')
    else:
        col2_num = col2 - 1
    
    # Read ALL data into memory at once using iter_rows (much faster than cell-by-cell)
    read_start = time.time()
    all_rows = list(sheet.iter_rows(min_row=start_row, values_only=True))
    read_time = time.time() - read_start
    print(f"⏱️  Read time: {read_time:.2f} seconds")
    
    # Perform concatenation in memory (Python list operations are very fast)
    concat_start = time.time()
    concatenated_values = []
    for row in all_rows:
        # Get values from both columns (row is 0-indexed tuple)
        val1 = row[col1_num] if col1_num < len(row) else None
        val2 = row[col2_num] if col2_num < len(row) else None
        
        # Handle None values - convert to empty string for concatenation
        val1_str = str(val1) if val1 is not None else ""
        val2_str = str(val2) if val2 is not None else ""
        
        # Concatenate with " - " separator
        concatenated = f"{val1_str} - {val2_str}" if val1_str or val2_str else ""
        concatenated_values.append(concatenated)
    concat_time = time.time() - concat_start
    print(f"⏱️  Concatenation time: {concat_time:.2f} seconds")
    
    # Insert a new column at column A (this shifts all existing columns to the right)
    insert_start = time.time()
    sheet.insert_cols(1)
    insert_time = time.time() - insert_start
    print(f"⏱️  Column insert time: {insert_time:.2f} seconds")
    
    # Write all concatenated values to column A at once
    write_start = time.time()
    for idx, value in enumerate(concatenated_values, start=start_row):
        sheet.cell(row=idx, column=1).value = value
    write_time = time.time() - write_start
    print(f"⏱️  Write time: {write_time:.2f} seconds")
    
    # Save the workbook
    save_start = time.time()
    output_path = save_path if save_path else file_path
    workbook.save(filename=output_path)
    save_time = time.time() - save_start
    print(f"⏱️  Save time: {save_time:.2f} seconds")
    
    total_time = time.time() - start_time
    print(f"✅ Successfully concatenated columns {col1} and {col2}, inserted into column A, and saved to {output_path}")
    print(f"⏱️  Total time: {total_time:.2f} seconds")
    
    return total_time


def mapToSlideCategories(file_path, sheet_name, col1, col2, start_row=1, save_path=None):
    """
    Concatenates two columns with " - " separator, maps the result through a dictionary,
    and inserts the mapped value into column A.
    All operations are performed in memory after loading for better performance.
    
    Parameters:
    -----------
    file_path : str
        Path to the Excel file
    sheet_name : str
        Name of the sheet to process
    col1 : str or int
        First column to concatenate (e.g., 'B' or 2)
    col2 : str or int
        Second column to concatenate (e.g., 'C' or 3)
    start_row : int, optional
        Row to start processing from (default: 1, header row is typically 1)
    save_path : str, optional
        Path to save the modified file. If None, overwrites the original file.
    
    Returns:
    --------
    float: Time taken in seconds
    """
    start_time = time.time()
    
    # Load the workbook
    load_start = time.time()
    workbook = load_workbook(filename=file_path, data_only=True)
    sheet = workbook[sheet_name]
    load_time = time.time() - load_start
    print(f"⏱️  Load time: {load_time:.2f} seconds")
    
    # Convert column letters to numbers if needed (0-indexed for list access)
    if isinstance(col1, str):
        col1_num = ord(col1.upper()) - ord('A')
    else:
        col1_num = col1 - 1
    
    if isinstance(col2, str):
        col2_num = ord(col2.upper()) - ord('A')
    else:
        col2_num = col2 - 1
    
    # Read ALL data into memory at once using iter_rows (much faster than cell-by-cell)
    read_start = time.time()
    all_rows = list(sheet.iter_rows(min_row=start_row, values_only=True))
    read_time = time.time() - read_start
    print(f"⏱️  Read time: {read_time:.2f} seconds")
    
    # Perform concatenation and dictionary mapping in memory
    concat_start = time.time()
    mapped_values = []
    for row in all_rows:
        # Get values from both columns (row is 0-indexed tuple)
        val1 = row[col1_num] if col1_num < len(row) else None
        val2 = row[col2_num] if col2_num < len(row) else None
        
        # Handle None values - convert to empty string for concatenation
        val1_str = str(val1) if val1 is not None else ""
        val2_str = str(val2) if val2 is not None else ""
        
        # Concatenate with " - " separator
        concatenated = f"{val1_str}{val2_str}" if val1_str or val2_str else ""
        
        # Map through dictionary - if key not found, output "NO SLIDE MATCH FOUND"
        mapped_value = slideCategories.get(concatenated, "NO SLIDE MATCH FOUND")
        mapped_values.append(mapped_value)
    concat_time = time.time() - concat_start
    print(f"⏱️  Concatenation and mapping time: {concat_time:.2f} seconds")
    
    # Insert a new column at column A (this shifts all existing columns to the right)
    insert_start = time.time()
    sheet.insert_cols(1)
    insert_time = time.time() - insert_start
    print(f"⏱️  Column insert time: {insert_time:.2f} seconds")
    
    # Write all mapped values to column A at once
    write_start = time.time()
    for idx, value in enumerate(mapped_values, start=start_row):
        sheet.cell(row=idx, column=1).value = value
    write_time = time.time() - write_start
    print(f"⏱️  Write time: {write_time:.2f} seconds")
    
    # Save the workbook
    save_start = time.time()
    output_path = save_path if save_path else file_path
    workbook.save(filename=output_path)
    save_time = time.time() - save_start
    print(f"⏱️  Save time: {save_time:.2f} seconds")
    
    total_time = time.time() - start_time
    print(f"✅ Successfully concatenated columns {col1} and {col2}, mapped through dictionary, inserted into column A, and saved to {output_path}")
    print(f"⏱️  Total time: {total_time:.2f} seconds")
    
    return total_time
